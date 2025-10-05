import openpyxl
from openpyxl.styles import Alignment

# Criar planilha
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Controle"

# Configurar cabeçalhos de datas
datas = ["27/08/25", "27/08/25", "27/08/25", "27/08/25"]
for i, data in enumerate(datas, start=2):  # Começa na coluna B
    ws.cell(row=1, column=i, value=data).alignment = Alignment(horizontal="center")

# Adicionar lista de itens (primeira coluna)
itens = ["ÁGUA", "COCA", "FANTA"]
for i, item in enumerate(itens, start=2):  # Começa na linha 2
    ws.cell(row=i, column=1, value=item)

# Ajustar largura das colunas
for col in range(1, len(datas) + 2):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

# Salvar planilha
caminho = r'C:\Users\cmsouza\Desktop\controle_bebidas.xlsx'
wb.save(caminho)

caminho
